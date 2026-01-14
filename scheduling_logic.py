import openpyxl  # For color import
import json
import io
import pandas as pd
from pandas import DataFrame, read_excel, isna, notna
#from collections import deque  
from datetime import datetime, timedelta
import os
import uuid

import firebase_manager as fm
fm.initialize_firebase()
ROLE_RULES = {}
GROUP_RULES = {"version": 1, "updated_at": None, "groups": []}
GROUP_RULES_FILE = "group_rules.json"

def initialize():
    """Initialize the module by loading data from files"""
    load_role_rules()
    load_group_rules()


def _now_utc_iso():
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def _default_group_rules():
    return {"version": 1, "updated_at": None, "groups": []}


def _normalize_group_rules(data):
    """
    Ensure group rules have a stable schema.

    Schema (version 1):
    {
      "version": 1,
      "updated_at": "2026-01-14T00:00:00Z" | null,
      "groups": [
        {
          "id": "uuid",
          "name": "韩娱组",
          "description": "",
          "active": true,
          "headcount_planned": 0 | null,
          "members": ["Alice", "Bob"],
          "requirements_windows": [
            {"day_type":"all|weekday|weekend","start":"00:00","end":"24:00","min_staff":1}
          ]
        }
      ]
    }
    """
    if not isinstance(data, dict):
        data = _default_group_rules()

    version = data.get("version", 1)
    groups = data.get("groups", [])
    if not isinstance(groups, list):
        groups = []

    norm_groups = []
    for g in groups:
        if not isinstance(g, dict):
            continue
        gid = g.get("id") or uuid.uuid4().hex
        name = str(g.get("name") or "").strip()
        if not name:
            # Skip unnamed groups to avoid UI/logic ambiguity
            continue
        members = g.get("members") or []
        if not isinstance(members, list):
            members = []
        members = [str(m).strip() for m in members if str(m).strip()]
        # de-dup while keeping order
        seen = set()
        members = [m for m in members if not (m in seen or seen.add(m))]

        windows = g.get("requirements_windows") or []
        if not isinstance(windows, list):
            windows = []
        norm_windows = []
        for w in windows:
            if not isinstance(w, dict):
                continue
            day_type = (w.get("day_type") or "all").strip().lower()
            if day_type not in {"all", "weekday", "weekend"}:
                day_type = "all"
            start = str(w.get("start") or "00:00").strip()
            end = str(w.get("end") or "24:00").strip()
            try:
                min_staff = int(w.get("min_staff", 1))
            except Exception:
                min_staff = 1
            if min_staff < 0:
                min_staff = 0
            norm_windows.append(
                {"day_type": day_type, "start": start, "end": end, "min_staff": min_staff}
            )

        norm_groups.append(
            {
                "id": gid,
                "name": name,
                "description": str(g.get("description") or ""),
                "active": bool(g.get("active", True)),
                "headcount_planned": g.get("headcount_planned", None),
                "members": members,
                "requirements_windows": norm_windows,
            }
        )

    return {"version": version, "updated_at": data.get("updated_at"), "groups": norm_groups}


def load_group_rules():
    """Load GROUP_RULES from Firebase if it exists, else fall back to local JSON file."""
    global GROUP_RULES
    data = fm.get_data("group_rules")
    if not data and os.path.exists(GROUP_RULES_FILE):
        try:
            with open(GROUP_RULES_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None

    GROUP_RULES = _normalize_group_rules(data)
    return GROUP_RULES


def save_group_rules(group_rules=None, also_write_local=True):
    """Save GROUP_RULES to Firebase (and optionally to local JSON seed file)."""
    global GROUP_RULES
    if group_rules is not None:
        GROUP_RULES = _normalize_group_rules(group_rules)

    GROUP_RULES["updated_at"] = _now_utc_iso()
    fm.save_data("group_rules", GROUP_RULES)

    # Also keep a copy in Firebase Storage for easy auditing / backup
    try:
        fm.save_json_to_storage("config/group_rules.json", GROUP_RULES)
    except Exception:
        # Storage sync is best-effort; DB is source of truth.
        pass

    if also_write_local:
        try:
            with open(GROUP_RULES_FILE, "w", encoding="utf-8") as f:
                json.dump(GROUP_RULES, f, ensure_ascii=False, indent=2)
        except Exception:
            # Local file write is best-effort; Firebase is source of truth in production.
            pass


def _sync_groups_after_employee_rename(old_name, new_name):
    """Update GROUP_RULES membership when an employee name changes."""
    global GROUP_RULES
    if not old_name or not new_name or old_name == new_name:
        return
    changed = False
    for g in GROUP_RULES.get("groups", []):
        members = g.get("members", [])
        if old_name in members:
            g["members"] = [new_name if m == old_name else m for m in members]
            # de-dup
            seen = set()
            g["members"] = [m for m in g["members"] if not (m in seen or seen.add(m))]
            changed = True
    if changed:
        save_group_rules(GROUP_RULES)


def _sync_groups_after_employee_delete(name):
    """Remove deleted employee name from all groups."""
    global GROUP_RULES
    if not name:
        return
    changed = False
    for g in GROUP_RULES.get("groups", []):
        members = g.get("members", [])
        if name in members:
            g["members"] = [m for m in members if m != name]
            changed = True
    if changed:
        save_group_rules(GROUP_RULES)

class Employee:
    def __init__(self, name, employee_type, additional_roles=None, start_time=None, end_time=None):
        self.name = name
        self.employee_type = employee_type  # Primary role (Role1)
        self.additional_roles = additional_roles or []  # List of additional roles
        self.start_time = start_time
        self.end_time = end_time

    def get_available_shifts(self):
        # Use primary role for shift determination
        if self.employee_type in ROLE_RULES:
            rule = ROLE_RULES[self.employee_type]
            if rule["rule_type"] == "shift_based":
                weekday_shifts = list(rule["shifts"]["weekday"].values())
                weekend_shifts = list(rule["shifts"]["weekend"].values())
                return list(set(weekday_shifts + weekend_shifts))
            elif rule["rule_type"] == "fixed_time":
                if self.start_time and self.end_time:
                    return [f"{self.start_time}-{self.end_time}"]
                return [rule.get("default_shift", "Shift Not Set")]
        return []
    
    def get_all_roles(self):
        """Return all roles (primary + additional)"""
        return [self.employee_type] + self.additional_roles


        


class Freelancer(Employee):
    def __init__(self, name):
        super().__init__(name, "Freelancer")
        
    def get_available_shifts(self):
        # Get current day of week to determine if it's a weekday or weekend
        today = datetime.now().weekday()
        day_type = "weekday" if today < 5 else "weekend"
        
        # Return all shifts for the current day type
        return list(ROLE_RULES["Freelancer"]["shifts"][day_type].values()) 

class SeniorEditor(Employee):
    def __init__(self, name):
        super().__init__(name, "SeniorEditor")
    
    def get_available_shifts(self):
        return ["13-22"]
class economics(Employee):
    def __init__(self, name):
        super().__init__(name, "economics")
    
    def get_available_shifts(self):
        return ["10-19"]
class Entertainment(Employee):
    def __init__(self, name):
        super().__init__(name, "Entertainment")
    
    def get_available_shifts(self):
        return ["10-19"]
class KoreanEntertainment(Employee):
    def __init__(self, name):
        super().__init__(name, "KoreanEntertainment")
    def get_available_shifts(self):
        return ["10-19"]

def init_employees():
    employees_raw = fm.get_data('employees')
    employees = []
    if employees_raw is None:
        return []
    for emp in employees_raw:
        if emp['role'] == 'Freelancer':
            employees.append(Freelancer(emp['name']))
        elif emp['role'] == 'SeniorEditor':
            employees.append(SeniorEditor(emp['name']))
        else:
            employees.append(Employee(emp['name'], emp['role']))
    return employees


def init_availability(start_date, employees):
    # Find the previous Sunday to start the calendar
    days_since_sunday = start_date.weekday() + 1  # +1 because Python's weekday() has Monday as 0
    first_sunday = start_date - timedelta(days=days_since_sunday % 7)
    
    # Create 4 weeks (28 days) of availability starting from the first Sunday
    return {
        (first_sunday + timedelta(days=i)).strftime("%Y-%m-%d"): {
            employee.name: [] for employee in employees
        } for i in range(28)  # 4 weeks
    }


def save_data(data):
    try:
        fm.save_data('availability', data)
    except Exception as e:
        raise RuntimeError(f"Firebase write failed: {str(e)}")

def load_data():
    data = fm.get_data('availability')
    return data if data else None


# Constants
EMPLOYEES = init_employees()
FREELANCERS = [employee.name for employee in EMPLOYEES if isinstance(employee, Freelancer)]

# New centralized role-based rules storage
# ROLE_RULES = {
#     "Freelancer": {
#         "rule_type": "shift_based",
#         "shifts": {
#             "weekday": {"early": "7-16", "day": "0930-1830", "night": "15-24"},
#             "weekend": {"early": "7-16", "day": "10-19", "night": "15-24"}
#         },
#         "requirements": {
#             "weekday": {"early": 1, "day": 1, "night": 2},
#             "weekend": {"early": 1, "day": 1, "night": 1}
#         }
#     },
#     "SeniorEditor": {
#         "rule_type": "fixed_time",
#         "default_shift": "13-22",
#     },
#     # Other roles can be added here with their specific rules
#     "economics": {
#         "rule_type": "fixed_time",
#         "default_shift": "10-19",
#     },
#     "Entertainment": {
#         "rule_type": "fixed_time",
#         "default_shift": "10-19",
#     },
#     "KoreanEntertainment": {
#         "rule_type": "fixed_time",
#         "default_shift": "10-19",
#     }
# }

def load_employees():
    data = fm.get_data('employees')
    if not data:
        return init_employees()
    # If your data is a dict, use `data.values()`; if list, use as is
    employee_list = data.values() if isinstance(data, dict) else data
    return [
        Employee(
            emp["name"],
            emp["role"],
            emp.get("additional_roles", []),
            emp.get("start_time"),
            emp.get("end_time")
        )
        for emp in employee_list
    ]

def import_employees_from_main_excel(excel_file, current_employees, addemployee_callback):
    """
    Imports employee availability with color codes from an Excel file.
    - Reads employee names from first row (ignores first column if it's date).
    - Reads dates from first column (ignores first row).
    - Reads cell values and cell colors.
    - Returns: detected names, missing names, and per-date availability dict including color info.
    :param excel_file: Uploaded file obj or file path (.xlsx)
    :param current_employees: list of employee names (str)
    :param addemployee_callback: function for UI/input (streamlit), expects (name, role, start, end)
    :return: names_detected, names_missing, availability_dict
    """

    # Use custom color-data extraction function placed in scheduling_logic.py
    employee_names, dates, data, color_matrix = get_excel_data_with_colors(excel_file)
    # Strip/clean names then check missing vs current
    names_detected = [str(x).strip() for x in employee_names if x and str(x).strip()]
    names_missing = [name for name in names_detected if name not in current_employees]

    # [Patch] Call `addemployee_callback` for each missing name (UI can handle prompts/inputs)
    for name in names_missing:
        addemployee_callback(name)
        # You may gather role/start/end via the callback, if needed

    # Build availability dict for all dates/employees, with color info for each cell
    availability_dict = {}
    for date_idx, date in enumerate(dates):
        # Normalize date key to ISO (YYYY-MM-DD) for downstream scheduling/validation.
        try:
            date_str = pd.to_datetime(date).strftime("%Y-%m-%d")
        except Exception:
            date_str = str(date)
        availability_dict[date_str] = {}
        for emp_idx, emp in enumerate(employee_names):
            cell_val = data[date_idx][emp_idx]
            cell_color = color_matrix[date_idx][emp_idx]
            availability_dict[date_str][emp] = {'value': cell_val, 'color': cell_color}

    # Return detected names, missing names, and detailed availability
    return names_detected, names_missing, availability_dict




def get_excel_data_with_colors(file):
    # Accept Streamlit-uploaded file or file path
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    
    # Read headers (assumes first row contains employee names, first column is dates)
    header_row = [cell.value for cell in ws[1]]
    dates = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    
    data = []
    colors = []

    for i in range(2, ws.max_row + 1):
        row_data = []
        row_colors = []
        for j in range(2, ws.max_column + 1):
            cell_val = ws.cell(row=i, column=j).value
            cell_color = ws.cell(row=i, column=j).fill.fgColor.rgb if ws.cell(row=i, column=j).fill.fgColor.type == 'rgb' else None
            row_data.append(cell_val)
            row_colors.append(cell_color)
        data.append(row_data)
        colors.append(row_colors)
    return header_row[1:], dates, data, colors


def sync_availability():
    employees = load_employees()
    availability = load_data()
    
    # Get current employee names
    current_employees = {e.name for e in employees}  # Use e.name
    if not availability:
        return
    # Update availability for each date
    for date in availability:
        # Remove deleted employees
        for emp_name in list(availability[date].keys()):
            if emp_name not in current_employees:
                del availability[date][emp_name]
                
        # Add new employees
        for emp in employees:
            if emp.name not in availability[date]:  # Use emp.name
                availability[date][emp.name] = []  # Use emp.name
    
    save_data(availability)

def validate_synchronization():
    employees = load_employees()
    availability = load_data()
    
    assert len(employees) > 0, "No employees found"
    
    # Check all employees exist in availability
    for emp in employees:
        display_name = emp.name
        for date in availability:
            assert display_name in availability[date], \
                f"{display_name} missing from {date}"
    
    # Check for orphaned availability entries
    all_displays = {emp.name for emp in employees}
    for date in availability:
        for emp_name in availability[date]:
            assert emp_name in all_displays, \
                f"Orphaned entry: {emp_name} on {date}"


def save_employees():
    # EMPLOYEES is assumed to be your global employee list
    json_data = [{
        "name": emp.name,
        "role": emp.employee_type,
        "additional_roles": emp.additional_roles,
        "start_time": emp.start_time,
        "end_time": emp.end_time
    } for emp in EMPLOYEES]
    fm.save_data('employees', json_data)



def add_employee(name, role, additional_roles=None, start_time=None, end_time=None):
    if role == 'Freelancer':
        new_emp = Freelancer(name)
        new_emp.additional_roles = additional_roles or []
    else:
        new_emp = Employee(name, role, additional_roles, start_time, end_time)
    
    EMPLOYEES.append(new_emp)
    save_employees()
    
    availability = load_data()
    
    if availability is None:
        availability = init_availability(datetime.now(), [new_emp])
    else:
        for date in availability:
            if role != 'Freelancer' and start_time and end_time:
                availability[date][new_emp.name] = [f"{start_time}-{end_time}"]
            else:
                availability[date][new_emp.name] = []
    save_data(availability)
    
    return new_emp


def edit_employee(old_name, new_name, new_role, additional_roles=None, new_start_time=None, new_end_time=None):
    for emp in EMPLOYEES:
        if emp.name == old_name:
            emp.name = new_name
            emp.employee_type = new_role
            emp.additional_roles = additional_roles or []
            if new_role != 'Freelancer':
                emp.start_time = new_start_time
                emp.end_time = new_end_time
            break

    # Keep group membership consistent when names change
    _sync_groups_after_employee_rename(old_name, new_name)

    availability = load_data()
    for date in availability:
        if old_name in availability[date]:
            current_shifts = availability[date][old_name]
            # Preserve leaves and special codes
            leaves = [s for s in current_shifts if s in {"AL", "CL", "PH", "ON", "自由調配", "half off"}]
            
            if new_role != 'Freelancer' and new_start_time and new_end_time:
                new_shift = f"{new_start_time}-{new_end_time}"
                # Only update non-leave days
                availability[date][new_name] = leaves if leaves else [new_shift]
    
    save_data(availability)
    save_employees()
    sync_availability()


def delete_employee(name):
    global EMPLOYEES
    # Find the employee to delete
    employee_to_delete = next((emp for emp in EMPLOYEES if emp.name == name), None)

    if employee_to_delete:
        # Remove employee from EMPLOYEES list
        EMPLOYEES.remove(employee_to_delete)

        # Save the updated employee list to JSON
        save_employees()

        # Remove from group membership rules
        _sync_groups_after_employee_delete(name)

        # Update availability data
        availability = load_data()
        for date in availability:
            if employee_to_delete.name in availability[date]:  # Use employee_to_delete.name
                del availability[date][employee_to_delete.name]  # Use employee_to_delete.name
        save_data(availability)
    else:
        print(f"Employee with name {name} not found.")


_last_generated_schedule = []

def get_last_generated_schedule():
    global _last_generated_schedule
    return _last_generated_schedule


def _parse_time_to_minutes(t):
    """
    Accepts: "7", "07", "0930", "09:30", "24", "24:00".
    Returns minutes in [0, 1440].
    """
    s = str(t).strip()
    if not s:
        return None
    s = s.replace("：", ":")
    if ":" in s:
        hh, mm = s.split(":", 1)
        hh = hh.strip()
        mm = mm.strip()
        if not hh:
            return None
        if mm == "":
            mm = "0"
        try:
            h = int(hh)
            m = int(mm)
        except Exception:
            return None
        if h == 24 and m == 0:
            return 1440
        if h < 0 or h > 23 or m < 0 or m > 59:
            return None
        return h * 60 + m

    # Pure digits: could be "7", "07", "0930"
    if not s.isdigit():
        return None
    if len(s) <= 2:
        try:
            h = int(s)
        except Exception:
            return None
        if h == 24:
            return 1440
        if h < 0 or h > 23:
            return None
        return h * 60
    if len(s) == 4:
        try:
            h = int(s[:2])
            m = int(s[2:])
        except Exception:
            return None
        if h == 24 and m == 0:
            return 1440
        if h < 0 or h > 23 or m < 0 or m > 59:
            return None
        return h * 60 + m
    return None


def _parse_shift_range(shift):
    """
    Parse "start-end" where each side can be like "7", "0930", "09:30", "24".
    Returns (start_min, end_min) in minutes on the same day, end is clamped to 1440.
    Supports overnight shifts by clamping to end-of-day for validation purposes.
    """
    if not shift or "-" not in str(shift):
        return None
    s = str(shift).strip()
    parts = s.split("-", 1)
    if len(parts) != 2:
        return None
    a = _parse_time_to_minutes(parts[0].strip())
    b = _parse_time_to_minutes(parts[1].strip())
    if a is None or b is None:
        return None

    # Overnight handling: e.g. 23-7 => treat as 23:00-24:00 for current date validation
    if b < a:
        b = 1440
    if b > 1440:
        b = 1440
    return (a, b)


def validate_schedule_against_group_rules(schedule_list, group_rules=None):
    """
    Validate generated schedule vs group coverage windows.
    Returns warnings (list[str]). Does not modify schedule.
    """
    group_rules = _normalize_group_rules(group_rules if group_rules is not None else GROUP_RULES)
    warnings = []
    if not schedule_list:
        return warnings

    # Prepare per-day assignment map: iso_date -> {employee_name: shift_string}
    for entry in schedule_list:
        try:
            d = entry.get("Date")
            if not d:
                continue
            # entry["Date"] is dd/mm/YYYY in current generator
            date_obj = datetime.strptime(d, "%d/%m/%Y")
            iso = date_obj.strftime("%Y-%m-%d")
            day_type = "weekend" if date_obj.weekday() >= 5 else "weekday"
        except Exception:
            continue

        # Only shift-like values count as coverage (must contain "-")
        for g in group_rules.get("groups", []):
            if not g.get("active", True):
                continue
            members = g.get("members", [])
            if not members:
                continue
            windows = g.get("requirements_windows", [])
            if not windows:
                continue

            for w in windows:
                w_day = (w.get("day_type") or "all").lower()
                if w_day not in {"all", day_type}:
                    continue
                start_m = _parse_time_to_minutes(w.get("start", "00:00"))
                end_m = _parse_time_to_minutes(w.get("end", "24:00"))
                if start_m is None or end_m is None:
                    continue
                if end_m < start_m:
                    # Reject invalid window silently (UI should avoid this)
                    continue
                min_staff = int(w.get("min_staff", 0) or 0)
                if min_staff <= 0:
                    continue

                # Check each hour bucket that starts within [start_m, end_m)
                hour_start = (start_m // 60) * 60
                while hour_start < end_m:
                    hour_end = min(hour_start + 60, 1440)
                    # Only validate if this hour overlaps window
                    if hour_end <= start_m:
                        hour_start += 60
                        continue
                    if hour_start >= end_m:
                        break

                    staffed = 0
                    for m in members:
                        v = entry.get(m)
                        rng = _parse_shift_range(v) if v else None
                        if not rng:
                            continue
                        s_min, e_min = rng
                        # overlap test with [hour_start, hour_end)
                        if max(s_min, hour_start) < min(e_min, hour_end):
                            staffed += 1
                    if staffed < min_staff:
                        warnings.append(
                            f"小组规则不足：{iso} {hour_start//60:02d}:00-{hour_end//60:02d}:00，"
                            f"小组“{g['name']}”需要≥{min_staff}人，实际{staffed}人。"
                        )
                    hour_start += 60

    return warnings

def generate_schedule(availability, start_date, export_to_excel=True, file_path=None):
    global _last_generated_schedule
    warnings = []
    
    # Get all dates from availability
    date_strings = sorted(availability.keys())
    dates = [datetime.strptime(d, "%Y-%m-%d") for d in date_strings]
    
    # Create a dictionary to organize schedule entries by date
    schedule_by_date = {}
    for date in dates:
        date_str = date.strftime("%d/%m/%Y")
        schedule_by_date[date_str] = {"Date": date_str}
    
    # Generate schedules for fulltime employees first
    for role_type in ROLE_RULES:
        if role_type != "Freelancer":  # Process all non-freelancer roles
            role_warnings = generate_fulltime_schedule_for_integrated(availability, dates, schedule_by_date, role_type)
            warnings.extend(role_warnings)
    
    # Generate freelancer schedule second
    freelancer_warnings = generate_freelancer_schedule_for_integrated(availability, dates, schedule_by_date)
    warnings.extend(freelancer_warnings)
    
    # Convert the date-organized dictionary to a flat schedule list
    schedule = [entry for _, entry in sorted(schedule_by_date.items())]

    # Validate vs custom group rules (if any)
    try:
        warnings.extend(validate_schedule_against_group_rules(schedule, GROUP_RULES))
    except Exception:
        # Validation should never break schedule generation
        pass
    
    # Store the generated schedule
    _last_generated_schedule = schedule
    
    # Export schedule to Excel if requested with the provided file path
    if export_to_excel and file_path:
        df = DataFrame(schedule)
        df.to_excel(file_path, index=False)
    
    return warnings

def generate_fulltime_schedule_for_integrated(availability, dates, schedule_by_date, role_type):
    """
    Generates schedules for fulltime employees of a specific role type and integrates them into the schedule_by_date dictionary.
    """
    warnings = []
    
    role_rules = ROLE_RULES[role_type]
    default_shift = role_rules.get("default_shift")
    
    employees = [emp.name for emp in EMPLOYEES if emp.employee_type == role_type]
    
    for date in dates:
        date_str = date.strftime("%d/%m/%Y")
        iso_date_str = date.strftime("%Y-%m-%d")
        
        # Assign shifts to employees based on actual availability
        for name in employees:
            # Check if the employee has availability data for this date
            if iso_date_str in availability and name in availability[iso_date_str]:
                employee_data = availability[iso_date_str][name]
                
                # Check if there's any data for this employee on this date
                if employee_data and len(employee_data) > 0:
                    # Check for leave types or custom shifts
                    first_entry = employee_data[0]
                    leave_types = ["AL", "CL", "PH", "ON", "自由調配", "half off"]
                    
                    if first_entry in leave_types:
                        # This is a leave entry
                        schedule_by_date[date_str][name] = first_entry
                    elif "-" in first_entry:
                        # This is a custom shift time
                        schedule_by_date[date_str][name] = first_entry
                    else:
                        # Unknown format, use default
                        schedule_by_date[date_str][name] = default_shift
                else:
                    # Empty array means unavailable
                    schedule_by_date[date_str][name] = "off"
            else:
                # If no data exists, use default shift
                schedule_by_date[date_str][name] = default_shift
    
    return warnings

def generate_freelancer_schedule_for_integrated(availability, dates, schedule_by_date):
    """
    Generates schedules for freelancers and integrates them into the schedule_by_date dictionary.
    """
    warnings = []
    
    freelancer_rules = ROLE_RULES["Freelancer"]
    shift_counts = {name: {"early": 0, "day": 0, "night": 0} for name in FREELANCERS}
    
    for date in dates:
        date_str = date.strftime("%d/%m/%Y")
        iso_date_str = date.strftime("%Y-%m-%d")
        day_type = 'weekend' if date.weekday() >= 5 else 'weekday'
        assigned_shifts = {name: 'off' for name in FREELANCERS}
        
        shifts = freelancer_rules["shifts"][day_type]
        shift_requirements = freelancer_rules["requirements"][day_type]
        
        # Process each shift type by priority
        for shift_name, required_count in sorted(shift_requirements.items(), key=lambda x: x[1], reverse=True):
            shift_time = shifts[shift_name]
            assigned_count = 0
            
            # Find available freelancers for this shift
            available_freelancers = []
            for name in FREELANCERS:
                if (iso_date_str in availability and 
                    name in availability[iso_date_str] and 
                    shift_time in availability[iso_date_str][name] and 
                    assigned_shifts[name] == 'off'):
                    weight = (1 / (len(availability[iso_date_str][name]) + 1)) + (1 / (shift_counts[name][shift_name] + 1))
                    available_freelancers.append((name, weight))
            
            # Sort freelancers by weight (higher weight = higher priority)
            available_freelancers.sort(key=lambda x: x[1], reverse=True)
            
            # Assign shifts
            for name, _ in available_freelancers:
                if assigned_count < required_count:
                    assigned_shifts[name] = shift_time
                    shift_counts[name][shift_name] += 1
                    assigned_count += 1
            
            # Check for understaffing
            if assigned_count < required_count:
                warnings.append(
                    f"Warning: {shift_name} shift on {date.strftime('%Y-%m-%d')} is understaffed. "
                    f"Required: {required_count}, Assigned: {assigned_count}."
                )
        
        # Add freelancer assignments to the schedule entry for this date
        for name, shift in assigned_shifts.items():
            schedule_by_date[date_str][name] = shift
    
    return warnings




def import_from_google_form(file_path):
    """
    Import employee availability data from Google Form responses Excel file.
    Handles both full-time and freelancer data formats.
    """
    try:
        df = read_excel(file_path)
        
        # Initialize availability data structure if not exists
        availability = load_data() or {}
        
        # Process each response row
        for _, row in df.iterrows():
            # Skip rows without name
            if isna(row.get('名字')):
                continue
                
            employee_name = row['名字']
            employee_type = row.get('請問您是全職還是兼職？')
            
            # Skip if employee type is not specified
            if isna(employee_type):
                continue
                
            # Process columns based on employee type
            if employee_type == '全職':
                # Process full-time employee leave options
                process_fulltime_availability(availability, row, employee_name)
            elif employee_type == '兼職':
                # Process freelancer shift selections
                process_freelancer_availability(availability, row, employee_name)
                
        # Save updated availability data
        save_data(availability)
        return "Google Form data imported successfully!"
    except Exception as e:
        raise ValueError(f"Failed to import Google Form data: {str(e)}")

def process_fulltime_availability(availability, row, employee_name):
    fulltime_cols = [col for col in row.index if col.startswith('全職 [') and ']' in col]
    
    # Find the employee object
    employee = next((emp for emp in EMPLOYEES if emp.name == employee_name), None)
    if not employee:
        return  # Skip if employee not found
        
    for col in fulltime_cols:
        date_str = col.split('[')[1].split(']')[0]
        date_parts = date_str.split('/')
        if len(date_parts) == 3:
            iso_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            
            if iso_date not in availability:
                availability[iso_date] = {}
            
            if employee_name not in availability[iso_date]:
                availability[iso_date][employee_name] = []
            
            leave_value = row[col]
            
            if notna(leave_value) and leave_value in ["AL", "CL", "PH", "ON", "自由調配", "half off"]:
                # This is a leave entry
                availability[iso_date][employee_name] = [leave_value]
            else:
                # This is a regular shift entry
                shift_value = row[col]
                if notna(shift_value) and "-" in shift_value:
                    # Update employee configuration with the shift from form
                    start_time, end_time = shift_value.split('-')
                    employee.start_time = start_time
                    employee.end_time = end_time
                    availability[iso_date][employee_name] = [shift_value]
                elif employee.employee_type in ROLE_RULES:
                    # Use employee's custom time if available, otherwise use default from role rules
                    if employee.start_time and employee.end_time:
                        availability[iso_date][employee_name] = [f"{employee.start_time}-{employee.end_time}"]
                    else:
                        rule = ROLE_RULES[employee.employee_type]
                        availability[iso_date][employee_name] = [rule["default_shift"]]
    
    # Save the updated employee configuration
    save_employees()




def process_freelancer_availability(availability, row, employee_name):
    """Process freelancer availability from form response."""
    # Identify freelancer date columns (format: '兼職 [DD/MM/YYYY]')
    freelancer_cols = [col for col in row.index if col.startswith('兼職 [') and ']' in col]
    
    for col in freelancer_cols:
        # Extract date from column name
        date_str = col.split('[')[1].split(']')[0]
        
        # Convert date format from DD/MM/YYYY to YYYY-MM-DD for internal storage
        date_parts = date_str.split('/')
        if len(date_parts) == 3:
            iso_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            
            # Create a datetime object to check if it's a weekday or weekend
            date_obj = datetime.strptime(iso_date, "%Y-%m-%d")
            is_weekend = date_obj.weekday() >= 5  # 5 and 6 are Saturday and Sunday
            
            # Initialize date in availability if not exists
            if iso_date not in availability:
                availability[iso_date] = {}
                
            # Initialize employee in date if not exists
            if employee_name not in availability[iso_date]:
                availability[iso_date][employee_name] = []
            
            # Get shift selections
            shift_value = row[col]
            
            # Skip if no shifts selected
            if isna(shift_value):
                continue
                
            # Process shift selections
            if shift_value == '全選':
                # All shifts selected
                availability[iso_date][employee_name] = ["7-16", "0930-1830" if not is_weekend else "10-19", "15-24"]
            else:
                # Parse individual shift selections
                shifts = []
                if '早更' in str(shift_value):
                    shifts.append("7-16")
                if '日更' in str(shift_value):
                    shifts.append("0930-1830" if not is_weekend else "10-17")
                if '夜更' in str(shift_value):
                    shifts.append("15-24")
                    
                availability[iso_date][employee_name] = shifts



def import_from_excel(file_path):
    df = read_excel(file_path)
    required_columns = {'Date', 'Employee', 'Shift'}
    if not required_columns.issubset(df.columns):
        raise ValueError(f"Excel file must contain columns: {required_columns}")
    
    availability = {}
    for _, row in df.iterrows():
        date_str = row['Date']
        employee_name = row['Employee']
        shift = row['Shift']
        
        if date_str not in availability:
            availability[date_str] = {name: [] for name in FREELANCERS}
        
        if employee_name not in availability[date_str]:
            availability[date_str][employee_name] = []
        
        if shift not in availability[date_str][employee_name]:
            availability[date_str][employee_name].append(shift)
        
        # Update employee configuration
        employee = next((emp for emp in EMPLOYEES if emp.name == employee_name), None)
        if employee and employee.employee_type != 'Freelancer':
            if '-' in shift and shift not in ["AL", "CL", "PH", "ON", "自由調配", "half off"]:
                start_time, end_time = shift.split('-')
                employee.start_time = start_time
                employee.end_time = end_time
    
    save_data(availability)
    save_employees()  # Save updated employee configurations
    return "Data imported successfully!"


def export_availability_to_excel(availability, file_path=None):
    """
    Exports availability data to an in-memory Excel file for downloading in Streamlit.
    """
    data = []
    for date, employees in availability.items():
        for employee_name, shifts in employees.items():
            # If shifts list is empty, create a row to indicate the employee exists
            if not shifts:
                data.append({"Date": date, "Employee": employee_name, "Shift": ""})
            else:
                for shift in shifts:
                    data.append({"Date": date, "Employee": employee_name, "Shift": shift})

    if not data:
        return io.BytesIO()

    df = pd.DataFrame(data) 
    
    # Create an in-memory buffer
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Availability')
    
    # Get the bytes value of the buffer
    processed_data = output.getvalue()
    return processed_data


def clear_availability(start_date, employees):
    # Reset custom times for all non-freelancer employees
    for emp in employees:
        if emp.employee_type != 'Freelancer':
            emp.start_time = None
            emp.end_time = None
    
    # Save updated employee configurations
    save_employees()
    
    # Initialize fresh availability data
    return init_availability(start_date, employees)

def add_role(role_name, role_config):
    """
    Add a new role type to ROLE_RULES
    
    Parameters:
    role_name (str): Name of the new role
    role_config (dict): Configuration for the role
    """
    global ROLE_RULES
    
    # Add the new role to ROLE_RULES
    ROLE_RULES[role_name] = role_config
    
    # Save the updated ROLE_RULES to a file
    save_role_rules()

def save_role_rules():
    """Save the ROLE_RULES dictionary to Firebase."""
    try:
        fm.save_data('role_rules', ROLE_RULES)
        try:
            fm.save_json_to_storage("config/role_rules.json", ROLE_RULES)
        except Exception:
            pass
    except Exception as e:
        print(f"Error saving role rules: {str(e)}")


def load_role_rules():
    """Load ROLE_RULES from Firebase if it exists."""
    global ROLE_RULES
    data = fm.get_data('role_rules')
    if data:
        ROLE_RULES = data
    # Optionally, else keep the default in memory if not present


initialize()    
    
    
